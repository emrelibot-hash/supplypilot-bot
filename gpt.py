# gpt.py
import os
import openai
import pandas as pd
import pdfplumber
from openpyxl import load_workbook

openai.api_key = os.getenv("OPENAI_API_KEY")


def extract_text_from_excel(file_path: str) -> str:
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    text = ""
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            line = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if line.strip():
                text += line + "\n"
    return text


def extract_text_from_pdf(file_path: str) -> str:
    text = ""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text


def ask_gpt_to_structure(text: str, is_boq=True) -> list[dict]:
    prompt = (
        "You are an assistant for structuring procurement data.\n"
        "Below is raw extracted text from a file. "
        "Please extract and return a structured JSON array with the following fields:\n"
    )

    if is_boq:
        prompt += "[No, Description, Unit, Qty]\n"
    else:
        prompt += "[No, Unit Price, Notes]\n"

    prompt += (
        "Ignore headers, currency symbols, and empty rows. "
        "Try to match formatting used in Excel or PDF. "
        "Respond ONLY with JSON array. No explanations.\n\n"
        f"Raw text:\n{text[:12000]}"
    )

    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        temperature=0,
        messages=[
            {"role": "system", "content": "You are a data extraction assistant."},
            {"role": "user", "content": prompt}
        ]
    )

    json_output = response.choices[0].message.content.strip()
    try:
        import json
        return json.loads(json_output)
    except Exception as e:
        print(f"[GPT Error] Failed to parse JSON: {e}")
        return []


def extract_boq_using_gpt(file_path: str) -> list[dict]:
    ext = file_path.lower()
    if ext.endswith(".pdf"):
        raw_text = extract_text_from_pdf(file_path)
    elif ext.endswith(".xls") or ext.endswith(".xlsx"):
        raw_text = extract_text_from_excel(file_path)
    else:
        raise ValueError("Unsupported BOQ file format")
    return ask_gpt_to_structure(raw_text, is_boq=True)


def extract_offer_using_gpt(file_path: str, supplier_name: str) -> dict:
    ext = file_path.lower()
    if ext.endswith(".pdf"):
        raw_text = extract_text_from_pdf(file_path)
    elif ext.endswith(".xls") or ext.endswith(".xlsx"):
        raw_text = extract_text_from_excel(file_path)
    else:
        raise ValueError("Unsupported offer file format")
    rows = ask_gpt_to_structure(raw_text, is_boq=False)
    return {
        "supplier": supplier_name,
        "rows": rows
    }


def translate_text(text: str, target_language="en") -> str:
    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            temperature=0,
            messages=[
                {"role": "system", "content": f"Translate to {target_language}."},
                {"role": "user", "content": text}
            ]
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"[GPT Error] Translation failed: {e}")
        return text
